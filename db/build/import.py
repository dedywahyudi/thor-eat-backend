import codecs
import pyodbc
import openpyxl
import os
import sys
import traceback
import json
import re
from pyodbc import DataError
from subprocess import call


if len(sys.argv) < 2:
    print 'No Excel file specified.'
    sys.exit(1)

xlsx_file = sys.argv[1]
file_path = '/usr/src/sql/excel/{}'.format(xlsx_file)
if not os.path.isfile(file_path):
    print 'The specified Excel file was not found in the expected path: {}'.format(file_path)
    sys.exit(1)

driver = '{ODBC Driver 17 for SQL Server}'
server = 'tcp:localhost'
database = 'ThorEarlyAlert'
username = 'sa'
password = 'Topcoder123'
cnxn = pyodbc.connect('DRIVER={}; SERVER={}; DATABASE={}; UID={}; PWD={}'.format(driver, server, database, username, password), autocommit=False)
cursor = cnxn.cursor()

ADMIN_USER_NAME = 'admin'
ADMIN_USER_ID = 1

subdivision_id_map = {}
product_line_id_map = {}

def cell_value(cell):
    if cell is not None and cell.value is not None:
        value = cell.value
        if (isinstance(value, basestring)):
            return value.strip()

        return value

    return ''

def clean_string(str):
    newStr = ''
    prev = ''
    strArr = [s.strip() for s in str.splitlines()]
    for idx, s in enumerate(strArr):
        s = s.strip()
        if not prev and not s:
            continue
        else:
            prev = s
            sClean = re.sub(r"^\W+", "", s)
            newStr += sClean
            if idx == len(strArr)-1:
                continue
            newStr += "\n"
    return newStr

# Insert organization and return the insert ID (just returns the ID if the organization exists in the DB)
def get_organization(name):
    # Check if the organization exists
    cursor.execute('SELECT Id FROM Organization WHERE Name = ?', [name])
    row = cursor.fetchone()
    if row:
        return row[0]

    cursor.execute('INSERT INTO Organization (Name) VALUES (?)', [name])
    #cnxn.commit()

    cursor.execute('SELECT @@IDENTITY')
    row = cursor.fetchone()
    if row:
        return row[0]

    return -1

def create_division(id, name):
    # Check if the division exists before inserting
    cursor.execute('SELECT Name FROM Division WHERE Id = ?', [id])
    row = cursor.fetchone()
    if not row:
        cursor.execute('INSERT INTO Division (Id, Name) VALUES (?, ?)', [id, name])

def get_division_id(name):
    cursor.execute('SELECT Id FROM Division WHERE Name = ?', [name])
    row = cursor.fetchone()
    if row:
        return row[0]

    return -1

def get_division_id_by_id(id):
    cursor.execute('SELECT Id FROM Division WHERE Id = ?', [id])
    row = cursor.fetchone()
    if row:
        return row[0]

    return -1

def get_subdivision(division_id, name):
    key = '{}:{}'.format(division_id, name)
    idx = 0
    subDivId = '{}{}'.format(division_id, idx)
    div_id = get_division_id_by_id(subDivId)
    while div_id > -1:
        idx = idx + 1
        subDivId = '{}{}'.format(division_id, idx)
        div_id = get_division_id_by_id(subDivId)

    if subdivision_id_map.has_key(key):
        return subdivision_id_map[key]

    # Check if the subdivision exists
    cursor.execute('SELECT Id FROM Division WHERE parentDivisionId = ? AND Name = ?', [division_id, name])
    row = cursor.fetchone()
    if row:
        subdivision_id = row[0]
        subdivision_id_map[key] = subdivision_id;
        return subdivision_id
    cursor.execute('INSERT INTO Division (Id, Name, region, parentDivisionId) VALUES (?, ?, ?, ?)', [subDivId, name, name, division_id])

    return subDivId

def insert_product_line(division_id, subdivision_id, name):
    sql = 'INSERT INTO ProductLine (DivisionId, Name, SubDivisionId) VALUES (?, ?, ?)' if subdivision_id > -1 \
        else 'INSERT INTO ProductLine (DivisionId, Name) VALUES (?, ?)'
    params = [division_id, name, subdivision_id] if subdivision_id > -1 else [division_id, name]
    cursor.execute(sql, params)

    cursor.execute('SELECT @@IDENTITY')
    row = cursor.fetchone()
    if row:
        product_line_id = row[0]
        # product_line_id_map[key] = product_line_id
        return product_line_id

    return -1

def get_product_line(division_id, subdivision_id, name):
    key = '{}:{}:{}'.format(division_id, subdivision_id, name) if subdivision_id > -1 else '{}:{}'.format(division_id, name)
    if product_line_id_map.has_key(key):
        return product_line_id_map[key]

    # Check if the subdivision exists
    sql = 'SELECT Id FROM ProductLine WHERE DivisionId = ? AND Name = ? AND SubDivisionId = ?' if subdivision_id > -1 \
        else 'SELECT Id FROM ProductLine WHERE DivisionId = ? AND Name = ?'
    params = [division_id, name, subdivision_id] if subdivision_id > -1 else [division_id, name]
    cursor.execute(sql, params)
    row = cursor.fetchone()
    if row:
        product_line_id = row[0]
        product_line_id_map[key] = product_line_id
        return product_line_id

    insert_product_line(division_id, subdivision_id, name)

def create_product_line_contact(product_line_id, email):
    # Check if the contact exists before inserting
    cursor.execute('SELECT Id FROM ProductLineContact WHERE ProductLineId = ? AND ContactEmail = ?', [product_line_id, email])
    row = cursor.fetchone()
    if not row:
        cursor.execute('INSERT INTO ProductLineContact (ProductLineId, ContactEmail) VALUES (?, ?)', [product_line_id, email])

def create_standard_division_contact(division_id, email):
    # Check if the contact exists before inserting
    cursor.execute('SELECT Id FROM StandardDivisionContact WHERE StandardDivisionId = ? AND ContactEmail = ?', [division_id, email])
    row = cursor.fetchone()
    if not row:
        cursor.execute('INSERT INTO StandardDivisionContact (StandardDivisionId, ContactEmail) VALUES (?, ?)', [division_id, email])


class Standard:
    @staticmethod
    def from_row(row_cells):
        s = Standard()
        s.name = cell_value(row_cells[1])
        s.organization = cell_value(row_cells[4])
        s.description = clean_string(cell_value(row_cells[5]))
        s.division_id = cell_value(row_cells[6])
        s.division = cell_value(row_cells[7])
        s.subdivision = cell_value(row_cells[8])
        s.product_lines = cell_value(row_cells[9]).split('\n')
        s.product_line_emails = cell_value(row_cells[10]).split('\n')
        s.critical_to_business = cell_value(row_cells[11])
        s.participant_email = cell_value(row_cells[12])
        s.comments = cell_value(row_cells[13])

        return s

    def insert(self):
        create_division(self.division_id, self.division)
        organization_id = get_organization(self.organization)
        subdivision_id = get_subdivision(self.division_id, self.subdivision) if len(self.subdivision) > 0 else -1
        product_line_id = 0
        # insert product line
        if len(self.product_lines) > 0:
            product_line_name = ','.join([x for x in self.product_lines if x.strip()])
            if len(product_line_name.strip()) > 0:
                product_line_id = insert_product_line(self.division_id, subdivision_id, product_line_name.strip())
                for email in self.product_line_emails:
                    emailCollection = re.findall('[^(@:\ ,)]+@[^@]+\.[^(@\\n\ ,;)]+', email)
                    print emailCollection
                    for contact_email in emailCollection:
                        if len(contact_email.strip()) > 0:
                            create_product_line_contact(product_line_id, contact_email.strip())

        # for product_line_name in self.product_lines:
        #     if len(product_line_name.strip()) > 0:
        #         product_line_id = get_product_line(self.division_id, subdivision_id, product_line_name)
        #         if (product_line_id > -1):
        #             product_line_ids.append(product_line_id)
        #             for email in self.product_line_emails:
        #                 if len(email.strip()) > 0:
        #                     create_product_line_contact(product_line_id, email)

        # insert the Standard
        cursor.execute('INSERT INTO Standard (OrganizationId, Name, Description, CreatedBy, CreatedDate) VALUES (?, ?, ?, ?, GETDATE())', [organization_id, self.name, self.description, 1])
        cursor.execute('SELECT @@IDENTITY')
        row = cursor.fetchone()

        if row:
            sql = 'INSERT INTO StandardDivision (StandardId, DivisionId, SubDivisionId, ProductLineId, StandardParticipant, CriticalToBusiness, Comment, IsApproved) ' \
                  'VALUES (?, ?, ?, ?, ?, ?, ?, \'true\')'
            # insert the standard divisions
            standard_id = row[0]
            is_critical = 'yes' == self.critical_to_business.lower()
            params = [standard_id, self.division_id, subdivision_id if subdivision_id > -1 else None, None, self.participant_email, is_critical, self.comments]

            # if len(product_line_ids) > 0:
            #     for id in product_line_ids:
            #         params[3] = id
            #         cursor.execute(sql, params)
            if product_line_id > 0:
                params[3] = product_line_id
                cursor.execute(sql, params)
            else:
                cursor.execute(sql, params)

            cursor.execute('SELECT @@IDENTITY')
            row = cursor.fetchone()
            if row:
                sd_id = row[0]
                if isinstance(self.participant_email, str) and len(self.participant_email.strip()) > 0:
                    create_standard_division_contact(sd_id, self.participant_email)

            # Finally, insert the history record
            cursor.execute('INSERT INTO History (RecordId, RecordName, Operation, RecordType, ModifiedDate) VALUES (?, ?, ?, ?, GETDATE())',
                           [standard_id, self.name, 'Insert', 'Standard'])


class CnSIssue:
    @staticmethod
    def from_row(row_cells):
        i = CnSIssue()
        i.division = cell_value(row_cells[1])
        i.description = clean_string(cell_value(row_cells[2]))
        i.action_plan = clean_string(cell_value(row_cells[3]))
        i.persons_responsible = cell_value(row_cells[4])
        i.estimated_date = cell_value(row_cells[5])
        i.created_datetime = cell_value(row_cells[6])
        i.engineer_review_date = cell_value(row_cells[8])
        i.engineering_leader = cell_value(row_cells[9])
        i.financial_impact = cell_value(row_cells[11])
        i.impact_summary = clean_string(cell_value(row_cells[12]))
        i.vp_gm_name = cell_value(row_cells[13])
        i.critical_role_leadership = cell_value(row_cells[14])
        i.succession_plan = cell_value(row_cells[15])
        i.offensive_defensive = cell_value(row_cells[16])
        i.external_partners = cell_value(row_cells[17])
        i.vp_gm_review_date = cell_value(row_cells[18])
        i.priority = cell_value(row_cells[19])
        i.region = cell_value(row_cells[20])
        i.roadmap_alignment = cell_value(row_cells[21])

        return i

    def insert(self):
        division_id = get_division_id(self.division)
        if division_id > -1:
            sql = 'INSERT INTO CnSIssue (DivisionId, Description, ActionPlan, PersonsResponsible, EstimatedCompletionDate, ' \
                  '    EngineerReviewDate, EngineerLeader, FinancialImpact, ImpactSummary, VPGMName, CriticalRoleLeadership, ' \
                  '    SuccessionPlan, OffensiveDefensive, ExternalPartners, VPGMReviewDate, Priority, Region, ' \
                  '    RoadmapAlignment, CreatedDate, [Read], CreatedBy) ' \
                   'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) '
            is_critical = 'yes' == self.critical_role_leadership.lower()
            params = [
                division_id, self.description, self.action_plan, self.persons_responsible, self.estimated_date,
                self.engineer_review_date, self.engineering_leader, self.financial_impact, self.impact_summary, self.vp_gm_name, is_critical,
                self.succession_plan, self.offensive_defensive, self.external_partners, self.vp_gm_review_date, self.priority, self.region,
                self.roadmap_alignment, self.created_datetime, False, ADMIN_USER_NAME
            ];
            cursor.execute(sql, params)
            cursor.execute('SELECT @@IDENTITY')
            row = cursor.fetchone()

            if row:
                cns_issue_id = row[0]
                cursor.execute(
                    'INSERT INTO History (RecordId, RecordName, Operation, RecordType, ModifiedDate, UserId) VALUES (?, ?, ?, ?, GETDATE(), ?)',
                    [cns_issue_id, None, 'Insert', 'CnSIssue', ADMIN_USER_ID])


# Open the workbook
wb = openpyxl.load_workbook(file_path, data_only = True, read_only = True)
sheet1 = wb[wb.sheetnames[0]] # The Standard sheet
sheet2 = wb[wb.sheetnames[1]] # The CnS Issue sheet
row_counter = 0
try:
    # Start from row 3 because row 1 contains "Confidential information" text and row 2 contains Column Names
    for row_cells in sheet1.iter_rows(min_row=3):
        std = Standard.from_row(row_cells)
        print json.dumps(std.__dict__)
        if len(std.name.strip()) == 0 or len(std.organization.strip()) == 0:
            continue

        try:
            std.insert()
            # row_counter += 1
            # if row_counter > 10:
            #     break
            try:
                print u'Inserted standard with name: {}'.format(std.name)
            except UnicodeEncodeError as e:
                # skip output errors with some of the strings'
                print u'Inserted standard successfully.'

        except DataError as e:
            u'Data error occurred for standard [{}]: {}'.format(std.name, e)

    for row_cells in sheet2.iter_rows(min_row=3):
        issue = CnSIssue.from_row(row_cells)
        if len(issue.division) == 0:
            continue

        try:
            issue.insert()
            print u'Inserted issue for division: {}'.format(issue.division)
        except DataError as e:
            u'Data error occurred for issue division [{}]: {}'.format(issue.division, e)

    cnxn.commit()
except Exception as e:
    print u'The data could not be imported into the database: {}'.format(e)
    print(traceback.format_exc())

print("/opt/mssql-tools/bin/sqlcmd -S localhost -U sa -P {} -d {} -i /usr/src/sql/backenddata.sql".format(password, database))
call(["/opt/mssql-tools/bin/sqlcmd", "-S", "localhost", "-U", "sa", "-P", password, "-d", database, "-i", "/usr/src/sql/backenddata.sql"])

